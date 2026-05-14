#!/usr/bin/env python3
"""Scan a 様式 xlsx for embedded instructions in label cells.

Usage:
    scan-form-instructions.py <form.xlsx>

行政・学術 様式 xlsx の label 行 (= 「研究業績等」 「研究目的」 等) には、
入力フォーマットの要件が **embedded instruction** として書かれていることが多い。
例: 「※著者（本人に下線）」 「半角数字で記入」 「200 字以内」 「箇条書き」 等。

label cell 全文を読まずに input cell を fill すると、 instruction を見落として
提出後 reject される (例: 2026-05-14 JST SPReAD で「研究業績欄の本人氏名
に下線」 を A14 内に書かれていたが見落とし、 複数回再提出で対応)。

本 script は xlsx の全 cell から **instruction keyword** にマッチする箇所を
抽出し、 keyword 別に group 化して表示する。 提出前に通して、 各 instruction
が input cell に反映されているかを user/Claude が手動 verify する。

Why script (not regex in convention doc):
- instruction の出現位置は form ごとに違う、 sheet も多岐 (= 1-5 枚目 + 留意事項)
- keyword は重複するが context (= 周辺文字列) で 「何への指示か」 が決まる
- script で keyword × cell をテーブル化 → user の cognitive load を軽減

Exit code 0 (= always informational、 critical/non-critical の判定はしない)。
"""

import sys
from openpyxl import load_workbook

# Instruction keywords with category labels.
# Each keyword → (category, hint about what to apply in input cells)
INSTRUCTION_PATTERNS = [
    # === 書式 (formatting) ===
    ("下線", "format", "input cell 内の対応箇所に rich text underline (CellRichText)"),
    ("本人に下線", "format", "input cell 内の本人氏名 (= 提出者の姓 or イニシャル) に rich text underline"),
    ("太字", "format", "input cell 内の対応箇所に rich text bold"),
    ("斜体", "format", "input cell 内の対応箇所に rich text italic"),

    # === 文字種 (charset) ===
    ("半角数字", "charset", "input cell は 半角 ASCII 数字のみ (= '1234'、 not '１２３４')"),
    ("半角英数字", "charset", "input cell は 半角 ASCII (= 0-9 a-z A-Z)"),
    ("全角", "charset", "input cell は 全角 (= 日本語の漢字・かな・全角英数記号)"),

    # === 量 (quantity) ===
    ("字以内", "limit", "input cell の文字数を制限内に収める (=LEN() で counter check)"),
    ("字程度", "limit", "input cell の文字数を目安値±10% 程度に収める"),
    ("文字以内", "limit", "input cell の文字数制限"),

    # === 構造 (structure) ===
    ("改行", "structure", "input cell 内で改行を使う (= '\\n' で区切る)"),
    ("箇条書き", "structure", "input cell を箇条書き形式 (= 各項目を改行で区切る)"),
    ("最大", "limit", "件数 / 文字数上限あり、 文中の数値を確認"),
    ("最小", "limit", "件数 / 文字数下限あり、 文中の数値を確認"),

    # === 添付・別紙 (attachment / separate) ===
    ("別紙", "attachment", "本 cell ではなく別紙 (= 別シート or 別 PDF) で提出する指示"),
    ("別添", "attachment", "別添ファイルとして提出する指示"),

    # === 認証 (authentication) ===
    ("署名", "auth", "本人署名 (= 手書き scan or signed PDF) 必須"),
    ("電子署名", "auth", "電子署名 (= 手書き scan 画像 or 認証付き電子署名)"),
    ("印鑑", "auth", "印鑑 (= 認印画像 or 朱印) 挿入"),
    ("認印", "auth", "認印 (= 朱印画像で代替可) 挿入"),
]


def scan(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=False)
    findings = {}  # keyword → [(sheet, coord, snippet)]

    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                for keyword, category, hint in INSTRUCTION_PATTERNS:
                    if keyword in v:
                        findings.setdefault((keyword, category, hint), []).append(
                            (sname, cell.coordinate, v)
                        )
    return findings


def _shorten(text, length=140):
    s = text.replace("\n", " ⏎ ")
    return f"{s[:length]}{'…' if len(s) > length else ''}"


def main():
    if len(sys.argv) != 2:
        print(__doc__, file=sys.stderr)
        sys.exit(2)

    findings = scan(sys.argv[1])
    if not findings:
        print("ℹ No instruction keywords matched. Form may be simple or keywords not yet listed.")
        sys.exit(0)

    # Group by category for readable output
    by_category = {}
    for (kw, cat, hint), occurrences in findings.items():
        by_category.setdefault(cat, []).append((kw, hint, occurrences))

    print(f"Found instruction keywords in {sum(len(occ) for v in by_category.values() for _,_,occ in v)} cell(s):\n")
    for cat in ("format", "charset", "limit", "structure", "attachment", "auth"):
        if cat not in by_category:
            continue
        items = by_category[cat]
        print(f"\n### [{cat}] {len(items)} keyword(s) hit")
        for kw, hint, occurrences in items:
            print(f"\n  ▶ 「{kw}」 → {hint}")
            for sname, coord, full in occurrences:
                print(f"      {sname}!{coord}: {_shorten(full)}")

    print("\n---")
    print("各 cell の input row (= label cell の 1 行下、 typically empty merged cell) で")
    print("instruction に従った format / charset / 字数 / 構造 / 添付 / 認証 が満たされて")
    print("いるかを手動 verify してください。")


if __name__ == "__main__":
    main()
