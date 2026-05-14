#!/usr/bin/env python3
"""Compare a filled-in 様式 xlsx against its template; flag label overwrites.

Usage:
    diff-form-xlsx.py <filled.xlsx> <template.xlsx>

Output categorizes each cell that differs between filled and template:
  ❌ LABEL_OVERWRITE: template has label-like text + filled has different text
                     (= the main bug class: applicant wrote narrative into a 様式 label row)
  ❌ LABEL_DELETED:   template has label-like text + filled is empty
  ℹ INPUT_FILLED:     template empty + filled has text (= normal applicant input)
  ℹ VALUE_CHANGED:    template has non-label content + filled differs (e.g. a placeholder)

Exit code 0 if no LABEL_* findings, 1 otherwise.

Why this script exists
----------------------
Japanese government 様式 xlsx forms typically structure each section as:
    Row N:   "<セクション>の必要性"      ← pre-printed LABEL (do not overwrite)
    Row N+1: (empty, merged across columns) ← applicant writes here
A common bug is writing the narrative INTO row N, overwriting the label.
2026-05-13 JST SPReAD application submitted with this exact bug in 3 places
(Sheet 3 rows 10/18/28); 教育研究支援課 flagged as "①様式の改変". The applicant
also reports having made the same mistake on prior form fills. Run this script
before submitting to catch the pattern mechanically.
"""

import sys
from openpyxl import load_workbook

# Heuristic: suffix patterns suggesting the cell is a 様式 label
LABEL_SUFFIXES = (
    "の必要性",   # SPReAD: "設備備品費、消耗品費の必要性"
    "の明細",     # SPReAD: "設備備品費の明細"
    "について",   # 学振等
    "の内容",     # 各種
    "の確認",     # 同意確認書系
    "の有無",     # 各種
    "の状況",     # 各種
)


def likely_label(text):
    if not isinstance(text, str):
        return False
    t = text.strip()
    return any(t.endswith(suf) for suf in LABEL_SUFFIXES)


def diff_workbook(filled_path, template_path):
    f = load_workbook(filled_path, data_only=False)
    t = load_workbook(template_path, data_only=False)
    common = [s for s in f.sheetnames if s in t.sheetnames]
    findings = {"LABEL_OVERWRITE": [], "LABEL_DELETED": [],
                "INPUT_FILLED": [], "VALUE_CHANGED": []}

    for sname in common:
        fs, ts = f[sname], t[sname]
        max_row = max(fs.max_row, ts.max_row)
        max_col = max(fs.max_column, ts.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                fv = fs.cell(r, c).value
                tv = ts.cell(r, c).value
                if fv == tv:
                    continue
                loc = f"{sname}!{fs.cell(r, c).coordinate}"
                tv_empty = tv is None or (isinstance(tv, str) and not tv.strip())
                fv_empty = fv is None or (isinstance(fv, str) and not fv.strip())
                if tv_empty:
                    findings["INPUT_FILLED"].append((loc, _shorten(fv)))
                elif fv_empty:
                    cls = "LABEL_DELETED" if likely_label(tv) else "VALUE_CHANGED"
                    findings[cls].append((loc, f"tpl={_shorten(tv)} -> empty"))
                else:
                    cls = "LABEL_OVERWRITE" if likely_label(tv) else "VALUE_CHANGED"
                    findings[cls].append((loc, f"tpl={_shorten(tv)} -> {_shorten(fv)}"))
    return findings


def _shorten(v, n=60):
    s = str(v).replace("\n", "⏎")
    return f"{s[:n]}{'…' if len(s) > n else ''}"


def main():
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        sys.exit(2)
    filled, template = sys.argv[1], sys.argv[2]
    findings = diff_workbook(filled, template)

    critical = 0
    for cls in ("LABEL_OVERWRITE", "LABEL_DELETED", "VALUE_CHANGED", "INPUT_FILLED"):
        items = findings[cls]
        if not items:
            continue
        is_critical = cls in ("LABEL_OVERWRITE", "LABEL_DELETED")
        marker = "❌" if is_critical else "ℹ"
        print(f"\n{marker} {cls} ({len(items)})")
        for loc, detail in items:
            print(f"  {loc}: {detail}")
        if is_critical:
            critical += len(items)

    if critical:
        print(f"\n❌ {critical} critical finding(s) — 様式 label cells were overwritten/deleted.")
        print("   Restore the original label text, and move your content to the next row down")
        print("   (the empty row immediately below the label is the intended input field).")
        sys.exit(1)
    print("\n✓ No label overwrites detected.")


if __name__ == "__main__":
    main()
